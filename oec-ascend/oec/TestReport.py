# encoding: utf-8

import pandas as pd
import os
import openpyxl
from oec.TestContext import TestContext
from oec.BaseTestCase import TestCase
from logging import getLogger

logger = getLogger("oec-ascend")


def gen_report(path: str, context: TestContext):
    log_dir = context.get_output_dir()
    src_path = os.path.join(path, "base_report.xlsx")
    path = os.path.join(log_dir, "report.xlsx")

    df = pd.read_excel(src_path, header=None)

    excel = openpyxl.load_workbook(src_path)
    sheet_name = excel.sheetnames[0]
    sheet = excel[sheet_name]
    for item in sheet.merged_cells:
        top_col, top_row, bottom_col, bottom_row = item.bounds
        df.iloc[top_row - 1 : bottom_row, top_col - 1 : bottom_col] = (
            item.start_cell.value
        )

    # 环境信息
    dfe = df.iloc[1:7, 3:]
    dfe.set_index([dfe.columns[0]], inplace=True)
    dfe.columns = pd.Index(["value"])
    info = context.infomation

    logger.debug(info)
    for key in dfe.index:
        dfe.loc[key, "value"] = info.get(key, "")
    logger.debug(dfe)
    # 测试结果
    dft = df.iloc[8:]
    dft.columns = dft.iloc[0]
    dft = dft.iloc[1:]

    # 初始化字典
    class info:
        def __init__(self):
            self.passed = 0
            self.tests:list[TestCase] = []

        def add(self, test: TestCase):
            self.tests.append(test)
            self.passed += test.count() if test.is_passed() else 0

        @property
        def total(self):
            return sum([test.count() if test.is_supported() else 0 for test in self.tests])
        
        @property
        def failed(self):
            return sum([test.count() if test.is_failed() else 0  for test in self.tests])
        
    dic = {}
    for i in range(len(dft)):
        dic.setdefault((dft.iat[i, 0], dft.iat[i, 1]), info())

    # 设置索引
    dft.set_index([dft.columns[0], dft.columns[1]], inplace=True)
    dft.sort_index()
    # 统计测试用例信息
    for _, test in context.get_used_tests().items():
        dic.setdefault(test.group, info())
        inf = dic[test.group]
        inf.add(test)

    details = pd.DataFrame(
        columns=["兼容性测试", "检测项", "用例编号", "测试内容", "结论"]
    )
    # 写入表格
    for k in dic:
        inf = dic[k]
        if inf.total != 0:
            dft.loc[k, "测试结果"] = f"{round(inf.passed/inf.total*100,2)}%"
            dft.loc[k, "结论"] = "PASS" if inf.failed == 0  else "FAILED"
        for test in inf.tests:
            g1, g2 = k
            details.loc[len(details)] = [
                g1,
                g2,
                test.name,
                test.get_test_content(),
                test.state.value,
            ]

    # 保存excel
    logger.debug(details)
    logger.debug(df)
    for i in range(1, 7):
        sheet.cell(i + 1, 5, df.iat[i, 4])

    for i in range(9, len(df)):
        for j in range(3, len(df.columns)):
            sheet.cell(i + 1, j + 1, df.iat[i, j])
    sheet2 = excel[excel.sheetnames[1]]

    for i in range(len(details)):
        for j in range(len(details.columns)):
            sheet2.cell(i + 2, j + 1, details.iat[i, j])
    excel.save(path)


if __name__ == "__main__":
    context = TestContext(".")
    gen_report("resource", context)
